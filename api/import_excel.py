"""POST /api/import_excel — Parse uploaded Excel file and append to Google Sheet.
Body: {"file_b64": "base64-encoded-xlsx"}
"""
from http.server import BaseHTTPRequestHandler
import base64, io
from api._shared import (
    json_response, read_body,
    existing_names_in_gsheet, gsheet_headers, gsheet_append_rows,
    require_auth, cors_headers,
)

COL_MAP = {
    "creator name": "Creator Name", "name": "Creator Name",
    "platform": "Platform", "platforms": "Platform",
    "tier": "Tier",
    "ccv": "CCV", "best ccv": "CCV", "avg ccv": "CCV",
    "country": "Country", "region": "Country",
    "content type": "Content Type", "category": "Content Type",
    "twitter": "Twitter Link", "twitter link": "Twitter Link",
    "instagram": "Instagram Link", "instagram link": "Instagram Link",
    "outreach status": "Outreach Status", "status": "Outreach Status",
    "notes": "Notes",
}

MAX_IMPORT_SIZE = 10 * 1024 * 1024  # 10 MB


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        if not require_auth(self, required_roles=("admin", "finance")):
            return
        try:
            body = read_body(self)
            file_b64 = body.get("file_b64", "")
            if not file_b64:
                json_response(self, 400, {"error": "no file_b64 provided"})
                return
            if len(file_b64) > MAX_IMPORT_SIZE:
                json_response(self, 400, {"error": "file too large (max 10 MB)"})
                return

            from openpyxl import load_workbook
            raw = base64.b64decode(file_b64)
            wb_in = load_workbook(filename=io.BytesIO(raw), read_only=True)
            ws_in = wb_in.active
            import_headers = [str(c.value).strip() if c.value else "" for c in next(ws_in.iter_rows(min_row=1, max_row=1))]

            # Map import columns to standard names
            col_idx = {}
            for i, h in enumerate(import_headers):
                mapped = COL_MAP.get(h.lower(), None)
                if mapped and mapped not in col_idx:
                    col_idx[mapped] = i

            if "Creator Name" not in col_idx:
                json_response(self, 400, {"error": "No 'Creator Name' or 'Name' column found"})
                return

            existing = existing_names_in_gsheet()
            sheet_headers = gsheet_headers()
            if not sheet_headers:
                json_response(self, 500, {"error": "Could not read Google Sheets headers"})
                return

            sheet_col_map = {h: i for i, h in enumerate(sheet_headers)}
            rows_to_append = []
            skipped = 0

            for row in ws_in.iter_rows(min_row=2, values_only=True):
                name_idx = col_idx.get("Creator Name")
                if name_idx is None or name_idx >= len(row):
                    continue
                name = str(row[name_idx] or "").strip()
                if not name or name.lower() in existing:
                    skipped += 1
                    continue

                # Build row aligned to sheet headers
                out_row = [""] * len(sheet_headers)
                for std_name, src_idx in col_idx.items():
                    val = str(row[src_idx] or "").strip() if src_idx < len(row) else ""
                    if val in ("nan", "None"):
                        val = ""
                    if std_name in sheet_col_map:
                        out_row[sheet_col_map[std_name]] = val

                rows_to_append.append(out_row)
                existing.add(name.lower())

            added = 0
            if rows_to_append:
                if gsheet_append_rows(rows_to_append):
                    added = len(rows_to_append)
                else:
                    skipped += len(rows_to_append)

            json_response(self, 200, {"ok": True, "added": added, "skipped": skipped})
        except Exception as e:
            json_response(self, 500, {"error": str(e)})

    def do_OPTIONS(self):
        self.send_response(204)
        cors_headers(self)
        self.end_headers()
