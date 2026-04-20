"""POST /api/export_excel — Generate a styled .xlsx from creator data and return it.
Body: {"rows": [{name, platform, handle, ccv, country, content, twitter, ...}, ...]}
"""
from http.server import BaseHTTPRequestHandler
import json, io, base64
from api._shared import read_body, require_auth, cors_headers


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        if not require_auth(self, required_roles=("admin", "finance")):
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            body = read_body(self)
            rows = body.get("rows") or []

            wb = Workbook()
            ws = wb.active
            ws.title = "Recast Export"

            headers = ["Creator Name", "Platform", "Handle", "CCV", "Country",
                       "Language", "Content Type", "Twitter", "Instagram", "Source", "Date"]

            # Header style
            header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F6EF7")
            thin_border = Border(
                left=Side(style="thin", color="D0D0D0"),
                right=Side(style="thin", color="D0D0D0"),
                top=Side(style="thin", color="D0D0D0"),
                bottom=Side(style="thin", color="D0D0D0"),
            )

            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for r_idx, row in enumerate(rows, 2):
                vals = [
                    row.get("name", ""), row.get("platform", ""), row.get("handle", ""),
                    int(row.get("ccv", 0) or 0), row.get("country", ""),
                    row.get("language", ""), row.get("content", ""),
                    row.get("twitter", ""), row.get("instagram", ""),
                    row.get("source", ""), row.get("date", ""),
                ]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=r_idx, column=c, value=v)
                    cell.border = thin_border
                    if c == 4:
                        cell.number_format = "#,##0"

            # Auto-width
            for col_cells in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 40)

            ws.freeze_panes = "A2"

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            data = buf.read()

            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", "attachment; filename=recast_export.xlsx")
            cors_headers(self)
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
        except Exception as e:
            body = json.dumps({"error": str(e)}).encode()
            self.send_response(500)
            self.send_header("Content-Type", "application/json")
            cors_headers(self)
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_response(204)
        cors_headers(self)
        self.end_headers()
